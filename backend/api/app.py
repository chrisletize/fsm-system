#!/usr/bin/env python3
"""
FSM Statement Generator - Web Interface
Simple Flask app for Michele to generate statements
"""

from flask import Flask, render_template, send_file, jsonify, request
from flask_cors import CORS
import psycopg2
import pandas as pd
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv
import openpyxl
from werkzeug.utils import secure_filename
import os
import sys
import uuid
import shutil
from datetime import datetime, date
from nc_tax_rates import get_tax_breakdown, get_county_rate_display
from branding import get_branding
from io import BytesIO
import zipfile
import re
from tax_processor import process_tax_report
import tempfile

# Add scripts directory to path so we can import our PDF generator
sys.path.append(os.path.join(os.path.dirname(__file__), '../../scripts'))
from generate_pdf_statement import generate_pdf_statement

load_dotenv()

app = Flask(__name__)
CORS(app)

# Upload configuration (ADD THIS HERE)
app.config['UPLOAD_FOLDER'] = '/tmp/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
def strip_excel_comments(input_path):
    """Remove corrupt comment XML from .xlsx files before openpyxl loads them."""
    temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(temp_fd)
    with zipfile.ZipFile(input_path, 'r') as zin:
        with zipfile.ZipFile(temp_path, 'w') as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if 'comments' in item.filename.lower() and item.filename.endswith('.xml'):
                    continue
                if item.filename.endswith('.rels') and 'worksheets' in item.filename:
                    text = data.decode('utf-8')
                    text = re.sub(
                        r'<Relationship[^>]*Target="[^"]*comments[^"]*"[^>]*/?>',
                        '', text, flags=re.IGNORECASE
                    )
                    data = text.encode('utf-8')
                if item.filename == '[Content_Types].xml':
                    text = data.decode('utf-8')
                    text = re.sub(
                        r'<Override[^>]*PartName="[^"]*comments[^"]*"[^>]*/?>',
                        '', text, flags=re.IGNORECASE
                    )
                    data = text.encode('utf-8')
                zout.writestr(item, data)
    return temp_path

def get_db_connection():
    return psycopg2.connect(
        host=os.getenv('DB_HOST'),
        port=os.getenv('DB_PORT'),
        database=os.getenv('DB_NAME'),
        user=os.getenv('DB_USER'),
        password=os.getenv('DB_PASSWORD'),
        cursor_factory=RealDictCursor
    )

@app.route('/')
def index():
    """Main page - show customers with outstanding balances"""
    return render_template('index.html')

@app.route('/api/customers')
def get_customers():
    """Get all customers with outstanding balances"""
    company_id = request.args.get('company_id', 2, type=int)  # Default to Get a Grip
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get customers with outstanding balances
    cur.execute("""
        SELECT 
            c.id,
            c.customer_name,
            c.account_number,
            c.contact_email,
            c.contact_phone,
            COUNT(i.id) as invoice_count,
            SUM(i.invoice_total_due) as total_due,
            MAX(i.invoice_date) as last_invoice_date,
            SUM(CASE WHEN (CURRENT_DATE - i.invoice_date) > 90 THEN i.invoice_total_due ELSE 0 END) as over_90_days,
            SUM(CASE WHEN (CURRENT_DATE - i.invoice_date) BETWEEN 61 AND 90 THEN i.invoice_total_due ELSE 0 END) as days_61_90,
            SUM(CASE WHEN (CURRENT_DATE - i.invoice_date) BETWEEN 31 AND 60 THEN i.invoice_total_due ELSE 0 END) as days_31_60,
            SUM(CASE WHEN (CURRENT_DATE - i.invoice_date) <= 30 THEN i.invoice_total_due ELSE 0 END) as current
        FROM customers c
        JOIN invoices i ON c.id = i.customer_id
        WHERE c.company_id = %s
          AND i.invoice_total_due > 0
        GROUP BY c.id, c.customer_name, c.account_number, c.contact_email, c.contact_phone
        ORDER BY total_due DESC
    """, (company_id,))
    
    customers = cur.fetchall()
    
    cur.close()
    conn.close()
    
    # Convert Decimal to float for JSON serialization
    for customer in customers:
        customer['total_due'] = float(customer['total_due'])
        customer['over_90_days'] = float(customer['over_90_days'])
        customer['days_61_90'] = float(customer['days_61_90'])
        customer['days_31_60'] = float(customer['days_31_60'])
        customer['current'] = float(customer['current'])
    
    return jsonify(customers)

@app.route('/api/companies')
def get_companies():
    """Get list of companies with branding"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM companies ORDER BY name")
    companies_data = cur.fetchall()
    cur.close()
    conn.close()
    
    # Add branding info to each company
    companies = []
    for company in companies_data:
        branding = get_branding(company['id'])
        companies.append({
            'id': company['id'],
            'name': company['name'],
            'logo_path': branding['logo'],
            'primary_color': branding['primary_color'],
            'secondary_color': branding['secondary_color']
        })
    
    return jsonify({'companies': companies})

@app.route('/api/branding/<int:company_id>')
def get_company_branding(company_id):
    """Get branding configuration for a company"""
    branding = get_branding(company_id)
    # Add /static/ prefix to logo path
    branding_with_url = branding.copy()
    branding_with_url['logo_url'] = f"/static/{branding['logo']}"
    return jsonify(branding_with_url)

@app.route('/api/generate-statement/<int:customer_id>')
def generate_statement(customer_id):
    """Generate PDF statement for a customer"""
    
    # Get customer name
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT customer_name, company_id FROM customers WHERE id = %s", (customer_id,))
    result = cur.fetchone()
    cur.close()
    conn.close()
    
    if not result:
        return jsonify({'error': 'Customer not found'}), 404
    
    customer_name = result['customer_name']
    customer_company_id = result['company_id']
    
    # Generate PDF
    output_dir = '/tmp/statements'
    os.makedirs(output_dir, exist_ok=True)
    
    safe_name = customer_name.replace(' ', '_').replace('/', '_')
    output_file = f"{output_dir}/statement_{safe_name}_{date.today().strftime('%Y%m%d')}.pdf"
    
    try:
        result = generate_pdf_statement(customer_name, output_file, customer_company_id)
        if result:
            return send_file(result, 
                           mimetype='application/pdf',
                           as_attachment=True,
                           download_name=f"statement_{safe_name}.pdf")
        else:
            return jsonify({'error': 'Failed to generate statement'}), 500

    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/api/summary')
def get_summary():
    """Get overall summary stats with aging buckets"""
    company_id = request.args.get('company_id', 2, type=int)
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get basic summary stats
    cur.execute("""
        SELECT
            COUNT(DISTINCT c.id) as customer_count,
            COUNT(i.id) as invoice_count,
            SUM(i.invoice_total_due) as total_due
        FROM customers c
        JOIN invoices i ON c.id = i.customer_id
        WHERE c.company_id = %s
          AND i.invoice_total_due > 0
    """, (company_id,))
    summary = cur.fetchone()
    
    # Calculate aging buckets
    cur.execute("""
        SELECT
            SUM(CASE 
                WHEN CURRENT_DATE - i.invoice_date <= 30 THEN i.invoice_total_due 
                ELSE 0 
            END) as current,
            SUM(CASE 
                WHEN CURRENT_DATE - i.invoice_date > 30 AND CURRENT_DATE - i.invoice_date <= 60 THEN i.invoice_total_due 
                ELSE 0 
            END) as days_30,
            SUM(CASE 
                WHEN CURRENT_DATE - i.invoice_date > 60 AND CURRENT_DATE - i.invoice_date <= 90 THEN i.invoice_total_due 
                ELSE 0 
            END) as days_60,
            SUM(CASE 
                WHEN CURRENT_DATE - i.invoice_date > 90 THEN i.invoice_total_due 
                ELSE 0 
            END) as days_90
        FROM invoices i
        WHERE i.company_id = %s
          AND i.invoice_total_due > 0
    """, (company_id,))
    aging = cur.fetchone()
    
    cur.close()
    conn.close()
    
    # Format response
    return jsonify({
        'customer_count': summary['customer_count'],
        'invoice_count': summary['invoice_count'],
        'total_due': float(summary['total_due']) if summary['total_due'] is not None else 0.0,
        'current': float(aging['current']) if aging['current'] is not None else 0.0,
        'days_30': float(aging['days_30']) if aging['days_30'] is not None else 0.0,
        'days_60': float(aging['days_60']) if aging['days_60'] is not None else 0.0,
        'days_90': float(aging['days_90']) if aging['days_90'] is not None else 0.0
    })
    
    cur.close()
    conn.close()
    
    summary['total_due'] = float(summary['total_due']) if summary['total_due'] is not None else 0.0
    
    return jsonify(summary)

@app.route('/upload', methods=['GET'])
def upload_page():
    """Show the upload interface"""
    return render_template('upload.html')

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Handle Excel file upload and import"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Only .xlsx and .xls files allowed'}), 400
    
    # Get selected company
    company_id = request.form.get('company_id')
    if not company_id:
        return jsonify({'error': 'No company selected'}), 400
    
    try:
        # Save file temporarily
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Parse and import
        result = import_servicefusion_excel(filepath, int(company_id))
        
        # Clean up
        os.remove(filepath)
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clear-company-data/<int:company_id>', methods=['DELETE'])
def clear_company_data(company_id):
    """Delete all invoices and customers for a company"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Delete invoices first (due to foreign key constraint)
        cur.execute("DELETE FROM invoices WHERE company_id = %s", (company_id,))
        invoices_deleted = cur.rowcount
        
        # Delete customers
        cur.execute("DELETE FROM customers WHERE company_id = %s", (company_id,))
        customers_deleted = cur.rowcount
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'invoices_deleted': invoices_deleted,
            'customers_deleted': customers_deleted
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def import_servicefusion_excel(filepath, company_id):
    """Parse ServiceFusion Excel export and import to database"""
    
    # Load workbook
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    
    # Row 6 is headers, data starts at row 7
    header_row = list(sheet.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    
    # Create column index mapping
    col_map = {header: idx for idx, header in enumerate(header_row)}
    
    stats = {
        'inserted': 0,
        'updated': 0,
        'skipped': 0,
        'errors': []
    }
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Process each invoice row
    for row_num, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        try:
            # Extract data
            invoice_number = str(row[col_map['Invoice#']])
            customer_name = row[col_map['Customer Name']]
            invoice_date = row[col_map['Invoice Date']]
            invoice_total = float(row[col_map['Invoice Total']] or 0)
            
            # AUTO-SPLIT: If uploading to Kleanit Charlotte (2) and customer has *FL*, route to Kleanit FL (4)
            target_company_id = company_id
            print(f"DEBUG: Processing '{customer_name}' - company_id={company_id}, has FL={'*FL*' in str(customer_name)}")
            if company_id == 1 and customer_name and '*FL*' in str(customer_name):
                target_company_id = 4  # Kleanit South Florida
                print(f"DEBUG: ✓ Routing '{customer_name}' to Kleanit FL")

            # Extract tax data
            tax_total = float(row[col_map.get('Tax Total')] or 0)
            tax_rate_name = row[col_map.get('Tax Rate Name')]
            
            invoice_total = float(row[col_map['Invoice Total']] or 0)
            amount_due = float(row[col_map['Invoice Total Due']] or 0)
            
            # Skip if no invoice number or already paid
            if not invoice_number:
                stats['skipped'] += 1
                continue
            
            # Check if customer exists
            cur.execute("""
                SELECT id FROM customers 
                WHERE company_id = %s AND customer_name = %s
            """, (target_company_id, customer_name))
            
            customer = cur.fetchone()
            
            if not customer:
                # Create customer
                cur.execute("""
                    INSERT INTO customers (company_id, customer_name, contact_email, contact_phone, 
                                         service_location_address_1, service_location_city, 
                                         service_location_state, service_location_zip)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    target_company_id,
                    customer_name,
                    row[col_map.get('Contact Email 1')],
                    row[col_map.get('Contact Phone 1')],
                    row[col_map.get('Service Location Address 1')],
                    row[col_map.get('Service Location City')],
                    row[col_map.get('Service Location State/Province')],
                    row[col_map.get('Service Location Zip/Post Code')]
                ))
                customer_id_val = cur.fetchone()['id']
            else:
                customer_id_val = customer['id']
            
            # Check if invoice exists
            cur.execute("""
                SELECT id FROM invoices 
                WHERE company_id = %s AND invoice_number = %s
            """, (target_company_id, invoice_number))
            
            existing = cur.fetchone()
            
            if existing:
                # Update existing invoice
                cur.execute("""
                    UPDATE invoices SET
                        customer_id = %s,
                        invoice_date = %s,
                        invoice_total = %s,
                        tax_total = %s,
                        tax_rate_name = %s,
                        invoice_total_due = %s,
                        invoice_status = %s
                    WHERE id = %s
                """, (customer_id_val, invoice_date, invoice_total, amount_due, 'Unpaid', existing['id']))
                stats['updated'] += 1
            else:
                # Insert new invoice
                cur.execute("""
                    INSERT INTO invoices (
                        company_id, customer_id, invoice_number,
                        invoice_date, invoice_status, invoice_total, invoice_total_due
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (target_company_id, customer_id_val, invoice_number,
                      invoice_date, 'Unpaid', invoice_total, amount_due))
                stats['inserted'] += 1
        
        except Exception as e:
            import traceback
            error_msg = traceback.format_exc()
            stats['errors'].append(f"Row {row_num}: {str(e)}")
            print(f"ERROR on row {row_num}: {error_msg}")  # This will show in Flask console
            continue
    
    conn.commit()
    cur.close()
    conn.close()
    
    return stats
# ========================================
# TAX REPORT API ENDPOINTS
# Add these to your existing app.py file
# ========================================

from datetime import datetime
import openpyxl
from io import BytesIO
from flask import send_file

# Route for tax report page
@app.route('/tax-report')
def tax_report():
    """Tax report page"""
    return render_template('tax-report.html')

# API: Get tax data for a company
@app.route('/api/tax-data/<int:company_id>')
def get_tax_data(company_id):
    """Get all tax transactions for a company with state/county/transit breakdown"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT 
            id,
            county,
            invoice_date,
            invoice_number,
            customer_name,
            job_number,
            total_sales,
            taxable_amount,
            tax_rate,
            tax_collected
        FROM tax_transactions
        WHERE company_id = %s
        ORDER BY county, invoice_date
    """, (company_id,))
    
    transactions = cur.fetchall()
    cur.close()
    conn.close()
    
    # Convert to list of dicts with Decimal to float AND add tax breakdown
    result = []
    for t in transactions:
        tax_collected = float(t['tax_collected'])
        breakdown = get_tax_breakdown(t['county'], tax_collected)
        
        result.append({
            'id': t['id'],
            'county': t['county'],
            'invoice_date': t['invoice_date'].isoformat() if t['invoice_date'] else None,
            'invoice_number': t['invoice_number'],
            'customer_name': t['customer_name'],
            'job_number': t['job_number'],
            'total_sales': float(t['total_sales']),
            'taxable_amount': float(t['taxable_amount']),
            'tax_rate': t['tax_rate'],
            'tax_collected': tax_collected,
            'state_tax': breakdown['state'],
            'county_tax': breakdown['county'],
            'transit_tax': breakdown['transit']
        })
    
    return jsonify(result)

# API: Upload tax report Excel file
@app.route('/api/upload-tax', methods=['POST'])
def upload_tax_report():
    """Upload and import ServiceFusion Tax Report"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    company_id = request.form.get('company_id')
    
    if not company_id:
        return jsonify({'error': 'No company selected'}), 400
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type'}), 400
    
    try:
        # Save file temporarily
        filename = secure_filename(file.filename)
        filepath = os.path.join('/tmp', filename)
        file.save(filepath)
        
        # Parse Excel file
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        inserted = 0
        updated = 0
        skipped = 0
        errors = []
        
        current_county = None
        
        # Start from row 2 (after headers)
        for row_num in range(2, ws.max_row + 1):
            try:
                # Get cell values
                col_a = ws.cell(row=row_num, column=1).value  # County name
                invoice_date = ws.cell(row=row_num, column=2).value
                invoice_number = ws.cell(row=row_num, column=3).value
                customer_name = ws.cell(row=row_num, column=4).value
                job_number = ws.cell(row=row_num, column=5).value
                total_sales = ws.cell(row=row_num, column=6).value
                taxable_amount = ws.cell(row=row_num, column=7).value
                tax_rate = ws.cell(row=row_num, column=8).value
                tax_collected = ws.cell(row=row_num, column=9).value
                
                # County name in column A
                if col_a and isinstance(col_a, str) and col_a.strip():
                    current_county = col_a.strip()
                
                # Skip empty rows
                if not invoice_number:
                    skipped += 1
                    continue
                
                # Parse invoice date
                if isinstance(invoice_date, datetime):
                    invoice_date_str = invoice_date.strftime('%Y-%m-%d')
                elif isinstance(invoice_date, str):
                    try:
                        parsed_date = datetime.strptime(invoice_date, '%m/%d/%Y')
                        invoice_date_str = parsed_date.strftime('%Y-%m-%d')
                    except:
                        invoice_date_str = invoice_date
                else:
                    skipped += 1
                    continue
                
                # Convert values
                total_sales_val = float(total_sales) if total_sales else 0
                taxable_amount_val = float(taxable_amount) if taxable_amount else 0
                tax_collected_val = float(tax_collected) if tax_collected else 0
                tax_rate_str = str(tax_rate) if tax_rate else '0%'
                
                # Check if transaction already exists
                cur.execute("""
                    SELECT id FROM tax_transactions 
                    WHERE company_id = %s AND invoice_number = %s
                """, (company_id, str(invoice_number)))
                
                existing = cur.fetchone()
                
                # Skip FL customers for Kleanit Charlotte
                if company_id == '1' and customer_name and '*FL*' in customer_name.upper():
                    skipped += 1
                    continue

                if existing:
                    # Update existing
                    cur.execute("""
                        UPDATE tax_transactions SET
                            county = %s,
                            invoice_date = %s,
                            customer_name = %s,
                            job_number = %s,
                            total_sales = %s,
                            taxable_amount = %s,
                            tax_rate = %s,
                            tax_collected = %s
                        WHERE id = %s
                    """, (current_county, invoice_date_str, customer_name or '',
                         job_number or '', total_sales_val, taxable_amount_val,
                         tax_rate_str, tax_collected_val, existing['id']))
                    updated += 1
                else:
                    # Insert new
                    cur.execute("""
                        INSERT INTO tax_transactions (
                            company_id, county, invoice_date, invoice_number,
                            customer_name, job_number, total_sales, taxable_amount,
                            tax_rate, tax_collected
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (company_id, current_county, invoice_date_str, str(invoice_number),
                         customer_name or '', job_number or '', total_sales_val,
                         taxable_amount_val, tax_rate_str, tax_collected_val))
                    inserted += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                continue
        
        conn.commit()
        cur.close()
        conn.close()
        
        # Clean up temp file
        os.remove(filepath)
        
        return jsonify({
            'inserted': inserted,
            'updated': updated,
            'skipped': skipped,
            'errors': errors
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API: Export tax data to Excel
@app.route('/api/export-tax/<int:company_id>')
def export_tax_data(company_id):
    """Export tax data to Excel"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get company name
    cur.execute("SELECT name FROM companies WHERE id = %s", (company_id,))
    company = cur.fetchone()
    company_name = company['name'] if company else 'Unknown'
    
    # Get all tax transactions
    cur.execute("""
        SELECT 
            county,
            invoice_date,
            invoice_number,
            customer_name,
            job_number,
            total_sales,
            taxable_amount,
            tax_rate,
            tax_collected
        FROM tax_transactions
        WHERE company_id = %s
        ORDER BY county, invoice_date
    """, (company_id,))
    
    transactions = cur.fetchall()
    cur.close()
    conn.close()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tax Report'
    
    # Headers
    headers = ['County', 'Invoice Date', 'Invoice #', 'Customer', 'Job #', 
               'Total Sales', 'Taxable Amount', 'Tax Rate', 'Tax Collected']
    ws.append(headers)
    
    # Data rows
    current_county = None
    county_totals = {}
    
    for t in transactions:
        county = t['county']
        
        # Track county totals
        if county not in county_totals:
            county_totals[county] = {
                'sales': 0,
                'taxable': 0,
                'tax': 0
            }
        
        county_totals[county]['sales'] += float(t['total_sales'])
        county_totals[county]['taxable'] += float(t['taxable_amount'])
        county_totals[county]['tax'] += float(t['tax_collected'])
        
        ws.append([
            county,
            t['invoice_date'].strftime('%m/%d/%Y') if t['invoice_date'] else '',
            t['invoice_number'],
            t['customer_name'],
            t['job_number'] or '',
            float(t['total_sales']),
            float(t['taxable_amount']),
            t['tax_rate'],
            float(t['tax_collected'])
        ])
    
    # Add summary sheet
    summary_ws = wb.create_sheet('Summary')
    summary_ws.append(['County', 'Total Sales', 'Taxable Amount', 'Tax Collected'])
    
    for county, totals in sorted(county_totals.items()):
        summary_ws.append([
            county,
            totals['sales'],
            totals['taxable'],
            totals['tax']
        ])
    
    # Grand total
    summary_ws.append([])
    summary_ws.append([
        'GRAND TOTAL',
        sum(t['sales'] for t in county_totals.values()),
        sum(t['taxable'] for t in county_totals.values()),
        sum(t['tax'] for t in county_totals.values())
    ])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    today = datetime.now().strftime('%Y-%m-%d')
    filename = f'TaxReport_{company_name.replace(" ", "_")}_{today}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# API: Clear tax data for a company
@app.route('/api/clear-tax-data/<int:company_id>', methods=['DELETE'])
def clear_tax_data(company_id):
    """Clear all tax transactions for a company"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("DELETE FROM tax_transactions WHERE company_id = %s", (company_id,))
        deleted_count = cur.rowcount
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'deleted': deleted_count,
            'message': f'Deleted {deleted_count} tax transactions'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ========================================
# BATCH STATEMENT GENERATION ENDPOINT
# Add this to your app.py file
# ========================================

import zipfile
from io import BytesIO
from datetime import datetime

# Add this import at the top if not already there:
# from flask import send_file

def clean_customer_name(name):
    """Convert customer name to proper Title Case with spaces"""
    # Replace underscores with spaces
    clean = name.replace('_', ' ')
    # Title case (capitalizes first letter of each word)
    clean = clean.title()
    return clean

@app.route('/api/generate-batch-statements', methods=['POST'])
def generate_batch_statements():
    """Generate PDF statements for multiple customers and return as ZIP file"""
    try:
        data = request.get_json()
        customer_ids = data.get('customer_ids', [])
        company_id = data.get('company_id')
        
        if not customer_ids:
            return jsonify({'error': 'No customers selected'}), 400
        
        if not company_id:
            return jsonify({'error': 'No company selected'}), 400
        
        # Import the PDF generation function
        sys.path.insert(0, '/home/chrisletize/fsm-system/scripts')
        from generate_pdf_statement import generate_pdf_statement
        
        # Create temp directory for PDFs
        temp_dir = f'/tmp/batch_statements_{uuid.uuid4().hex[:8]}'
        os.makedirs(temp_dir, exist_ok=True)
        
        # Create ZIP file in memory
        zip_buffer = BytesIO()
        
        # Get company name for filename
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT name FROM companies WHERE id = %s", (company_id,))
        company = cur.fetchone()
        company_name = company['name'] if company else 'Company'
        cur.close()
        conn.close()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Generate PDF for each customer
            for customer_id in customer_ids:
                try:
                    # Get customer name
                    conn = get_db_connection()
                    cur = conn.cursor()
                    cur.execute("SELECT customer_name, company_id FROM customers WHERE id = %s AND company_id = %s",
                               (customer_id, company_id))
                    customer = cur.fetchone()
                    cur.close()
                    conn.close()
                    
                    if not customer:
                        continue
                    
                    customer_name = customer['customer_name']
                    customer_company_id = customer['company_id']
                    
                    # Generate PDF to temp file
                    safe_name = customer_name.replace(' ', '_').replace('/', '_')
                    temp_pdf = f"{temp_dir}/statement_{safe_name}.pdf"
                    
                    # Generate the PDF using customer_id (more reliable than name matching)
                    result = generate_pdf_statement(
                        customer_name_search=None,
                        output_file=temp_pdf,
                        company_id=customer_company_id,
                        customer_id=customer_id
                    )
                    
                    if result and os.path.exists(result):
                        # Add PDF to ZIP
                        with open(result, 'rb') as pdf_file:
                            # Remove asterisks and extra spaces to avoid Windows ZIP preview issues
                            clean_name = clean_customer_name(customer_name).replace('*', '').replace('  ', ' ').strip()
                            zip_file.writestr(f'Statement - {clean_name}.pdf', pdf_file.read())
                        
                        # Clean up temp file
                        os.remove(result)
                    
                except Exception as e:
                    print(f"Error generating statement for customer {customer_id}: {e}")
                    continue
        
        # Clean up temp directory
        shutil.rmtree(temp_dir, ignore_errors=True)
        # Prepare ZIP for download
        zip_buffer.seek(0)
        
        # Prepare ZIP for download
        zip_buffer.seek(0)
        
        # Generate filename
        today = datetime.now().strftime('%Y-%m-%d')
        filename = f'Statements_{company_name.replace(" ", "_")}_{today}.zip'
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error in batch generation: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/process-tax-report', methods=['POST'])
def api_process_tax_report():
    """Process tax report and transaction report to create cash-basis tax breakdown"""
    try:
        tax_file = request.files.get('tax_report')
        transaction_file = request.files.get('transaction_report')
        company_id = request.form.get('company_id')

        if not all([tax_file, transaction_file, company_id]):
            return jsonify({'success': False, 'error': 'Missing required fields'})

        # Save uploaded files temporarily
        tax_path = os.path.join(tempfile.gettempdir(), f'tax_report_{company_id}.xlsx')
        transaction_path = os.path.join(tempfile.gettempdir(), f'transaction_report_{company_id}.xlsx')

        tax_file.save(tax_path)
        transaction_file.save(transaction_path)

        # Process the reports
        result = process_tax_report(tax_path, transaction_path, company_id)

        # Clean up temp files
        os.remove(tax_path)
        os.remove(transaction_path)

        return jsonify(result)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


# ========================================
# OUTLOOK INTEGRATION ENDPOINTS
# ========================================

@app.route('/api/prepare-outlook-email/<int:customer_id>', methods=['POST'])
def prepare_outlook_email(customer_id):
    """Generate PowerShell script + PDF for single customer email"""
    try:
        data = request.get_json()
        company_id = data.get('company_id')

        if not company_id:
            return jsonify({'error': 'No company selected'}), 400

        # Get customer info
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT
                c.customer_name,
                c.contact_email,
                co.name as company_name,
                SUM(i.invoice_total_due) as total_due
            FROM customers c
            JOIN companies co ON c.company_id = co.id
            LEFT JOIN invoices i ON c.id = i.customer_id AND i.invoice_total_due > 0
            WHERE c.id = %s AND c.company_id = %s
            GROUP BY c.id, c.customer_name, c.contact_email, co.name
        """, (customer_id, company_id))

        customer = cur.fetchone()
        cur.close()
        conn.close()

        if not customer:
            return jsonify({'error': 'Customer not found'}), 404

        if not customer['contact_email']:
            return jsonify({'error': 'Customer has no email address on file'}), 400

        customer_name = customer['customer_name']

        # Map to official company names
        company_name_map = {
            'Kleanit Charlotte': 'Kleanit',
            'Get a Grip Resurfacing of Charlotte': 'Get a Grip Resurfacing of Charlotte',
            'CTS of Raleigh': 'CTS of Raleigh',
            'Kleanit South Florida': 'Kleanit'
        }
        company_name = company_name_map.get(customer['company_name'], customer['company_name'])

        total_due = float(customer['total_due']) if customer['total_due'] else 0.0

        # Generate PDF statement
        temp_dir = f'/tmp/outlook_email_{uuid.uuid4().hex[:8]}'
        os.makedirs(temp_dir, exist_ok=True)

        safe_name = customer_name.replace(' ', '_').replace('/', '_')
        pdf_filename = f"Statement_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
        pdf_path = os.path.join(temp_dir, pdf_filename)

        # Generate the PDF
        result_pdf = generate_pdf_statement(customer_name, pdf_path, company_id)

        if not result_pdf or not os.path.exists(result_pdf):
            return jsonify({'error': 'Failed to generate PDF statement'}), 500

        # Import Outlook functions
        from outlook_integration import generate_individual_email_script, save_script_to_file

        # Generate PowerShell script
        script_content = generate_individual_email_script(
            customer_name=customer_name,
            customer_email=customer['contact_email'],
            company_name=company_name,
            total_due=total_due,
            pdf_filename=pdf_filename
        )

        script_filename = f"Email_{safe_name}.ps1"
        script_path = os.path.join(temp_dir, script_filename)
        save_script_to_file(script_content, script_path)

        # Create ZIP file
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            readme = f"""Outlook Email Package for {customer_name}
=====================================

Contents:
1. {script_filename} - PowerShell script
2. {pdf_filename} - Customer statement

Instructions:
1. Extract all files to a folder
2. Right-click {script_filename}
3. Select "Run with PowerShell"
4. Review the draft in Outlook
5. Click Send when ready

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
            zip_file.writestr('README.txt', readme)
            zip_file.write(script_path, script_filename)
            zip_file.write(result_pdf, pdf_filename)

        # Clean up
        os.remove(script_path)
        os.remove(result_pdf)
        shutil.rmtree(temp_dir, ignore_errors=True)
        zip_buffer.seek(0)

        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'Email_{safe_name}.zip'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/prepare-outlook-batch', methods=['POST'])
def prepare_outlook_batch():
    """Generate PowerShell script + PDFs for batch emails"""
    try:
        data = request.get_json()
        customer_ids = data.get('customer_ids', [])
        company_id = data.get('company_id')

        if not customer_ids:
            return jsonify({'error': 'No customers selected'}), 400

        if not company_id:
            return jsonify({'error': 'No company selected'}), 400

        # Get company name
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT name FROM companies WHERE id = %s", (company_id,))
        company = cur.fetchone()
        
        # Map to official company names
        company_name_map = {
            'Kleanit Charlotte': 'Kleanit',
            'Get a Grip Resurfacing of Charlotte': 'Get a Grip Resurfacing of Charlotte',
            'CTS of Raleigh': 'CTS of Raleigh',
            'Kleanit South Florida': 'Kleanit'
        }
        company_name = company_name_map.get(company['name'], company['name']) if company else 'Company'
        cur.close()
        conn.close()

        # Create temp directory
        temp_dir = f'/tmp/outlook_batch_{uuid.uuid4().hex[:8]}'
        os.makedirs(temp_dir, exist_ok=True)

        customers_data = []
        pdf_files = []

        # Generate PDFs and collect customer data
        for customer_id in customer_ids:
            try:
                conn = get_db_connection()
                cur = conn.cursor()
                cur.execute("""
                    SELECT
                        c.customer_name,
                        c.contact_email,
                        SUM(i.invoice_total_due) as total_due
                    FROM customers c
                    LEFT JOIN invoices i ON c.id = i.customer_id AND i.invoice_total_due > 0
                    WHERE c.id = %s AND c.company_id = %s
                    GROUP BY c.id, c.customer_name, c.contact_email
                """, (customer_id, company_id))

                customer = cur.fetchone()
                cur.close()
                conn.close()

                if not customer or not customer['contact_email']:
                    continue

                customer_name = customer['customer_name']
                total_due = float(customer['total_due']) if customer['total_due'] else 0.0

                clean_name = clean_customer_name(customer_name).replace('*', '').replace('  ', ' ').strip()
                pdf_filename = f"Statement - {clean_name}.pdf"
                pdf_path = os.path.join(temp_dir, pdf_filename)

                result_pdf = generate_pdf_statement(customer_name, pdf_path, company_id)

                if result_pdf and os.path.exists(result_pdf):
                    customers_data.append({
                        'name': customer_name,
                        'email': customer['contact_email'],
                        'total_due': total_due,
                        'pdf_filename': pdf_filename
                    })
                    pdf_files.append((result_pdf, pdf_filename))

            except Exception as e:
                print(f"Error processing customer {customer_id}: {e}")
                continue

        if not customers_data:
            return jsonify({'error': 'No valid customers to process'}), 400

        # Import Outlook functions
        from outlook_integration import generate_batch_email_script, save_script_to_file

        # Generate batch script
        script_content = generate_batch_email_script(
            customers_data=customers_data,
            company_name=company_name
        )

        script_filename = f"Batch_Email_{len(customers_data)}_Customers.ps1"
        script_path = os.path.join(temp_dir, script_filename)
        save_script_to_file(script_content, script_path)

        # Create ZIP
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            readme = f"""Outlook Batch Email Package
============================

Company: {company_name}
Customers: {len(customers_data)}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Instructions:
1. Extract all files
2. Right-click {script_filename}
3. Select "Run with PowerShell"
4. Review customer list and confirm
5. Open Outlook Drafts folder
6. Review and send each email

Customers:
"""
            for customer in customers_data:
                readme += f"  • {customer['name']} <{customer['email']}> - ${customer['total_due']:,.2f}\n"

            zip_file.writestr('README.txt', readme)
            zip_file.write(script_path, script_filename)

            for pdf_path, pdf_filename in pdf_files:
                zip_file.write(pdf_path, pdf_filename)

        # Clean up
        os.remove(script_path)
        for pdf_path, _ in pdf_files:
            try:
                os.remove(pdf_path)
            except:
                pass
        shutil.rmtree(temp_dir, ignore_errors=True)
        zip_buffer.seek(0)

        today = datetime.now().strftime('%Y-%m-%d')
        filename = f'Batch_Email_{company_name.replace(" ", "_")}_{len(customers_data)}_customers_{today}.zip'

        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/generate-tax-report-pdf', methods=['POST'])
def generate_tax_report_pdf():
    """Generate PDF from current tax report data"""
    try:
        data = request.get_json()
        company_id = data.get('company_id')
        report_data = data.get('report_data')
        
        if not company_id or not report_data:
            return jsonify({'error': 'Missing company_id or report_data'}), 400
        
        # Import PDF generator
        sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../scripts'))
        from generate_pdf_tax_report import generate_pdf_tax_report
        
        # Get company branding
        branding = get_branding(int(company_id))
        
        # Generate filename
        company_name = branding['name'].replace(' ', '_')
        today = datetime.now().strftime('%Y-%m-%d')
        output_file = f'/tmp/TaxReport_{company_name}_{today}.pdf'
        
        # Generate PDF
        result = generate_pdf_tax_report(report_data, branding, output_file)
        
        if result and os.path.exists(result):
            return send_file(
                result,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f'TaxReport_{company_name}_{today}.pdf'
            )
        else:
            return jsonify({'error': 'Failed to generate PDF'}), 500
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/recency-report')
def recency_report():
    """Customer Recency Report Generator page"""
    return render_template('recency_report.html')

@app.route('/api/recency/upload', methods=['POST'])
def upload_recency_data():
    """Upload and process ServiceFusion Customer Revenue Report"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        company_id = request.form.get('company_id')

        if not company_id:
            return jsonify({'error': 'No company selected'}), 400

        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        # Save uploaded file to temp location
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()
        # Strip corrupt comments from ServiceFusion Excel files
        cleaned_path = strip_excel_comments(temp_file.name)
        os.unlink(temp_file.name)  # Done with original

        # Read cleaned file with pandas (header is auto-detected as row 0)
        df = pd.read_excel(cleaned_path, engine='openpyxl')
        os.unlink(cleaned_path)
        
        # Verify required columns
        if 'Customer' not in df.columns or 'Date' not in df.columns:
            return jsonify({'error': 'Excel file missing required columns (Customer, Date)'}), 400

        # Process rows
        conn = get_db_connection()
        cur = conn.cursor()

        processed = 0
        inserted = 0
        updated = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                customer_name = row.get('Customer')
                date_value = row.get('Date')

                if pd.isna(customer_name) or pd.isna(date_value):
                    continue

                customer_name = str(customer_name).strip()

                # Parse date
                if isinstance(date_value, pd.Timestamp):
                    job_date = date_value.date()
                elif isinstance(date_value, datetime):
                    job_date = date_value.date()
                elif isinstance(date_value, str):
                    try:
                        job_date = datetime.strptime(date_value, '%m/%d/%Y').date()
                    except:
                        continue
                else:
                    continue

                # Find or create customer
                cur.execute("""
                    SELECT id FROM customers
                    WHERE customer_name = %s AND company_id = %s
                    LIMIT 1
                """, (customer_name, company_id))

                customer = cur.fetchone()

                if not customer:
                    # Create new customer
                    cur.execute("""
                        INSERT INTO customers (customer_name, company_id, created_at)
                        VALUES (%s, %s, NOW())
                        RETURNING id
                    """, (customer_name, company_id))
                    result = cur.fetchone()
                    customer_id_val = result['id'] if result else None
                    conn.commit()
                else:
                    customer_id_val = customer['id']

                # Insert or ignore job date (UPSERT)
                cur.execute("""
                    INSERT INTO customer_job_dates (customer_id, job_date, source, created_at, created_by)
                    VALUES (%s, %s, 'servicefusion_import', NOW(), 'recency_upload')
                    ON CONFLICT (customer_id, job_date) DO NOTHING
                    RETURNING id
                """, (customer_id_val, job_date))

                result = cur.fetchone()
                if result and result.get('id'):
                    inserted += 1
                else:
                    updated += 1

                processed += 1

                # Commit every 100 rows
                if processed % 100 == 0:
                    conn.commit()

            except Exception as e:
                import traceback
                error_detail = f"Row {idx}: {str(e)}\n{traceback.format_exc()}"
                errors.append(error_detail)
                print(f"ERROR: {error_detail}")
                continue

        # Final commit
        conn.commit()
        cur.close()
        conn.close()

        return jsonify({
            'success': True,
            'processed': processed,
            'inserted': inserted,
            'updated': updated,
            'errors': errors[:10]
        })

    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/api/recency/report', methods=['POST'])
def generate_recency_report():
    """Generate recency report for selected company"""
    try:
        data = request.get_json()
        company_id = data.get('company_id')
        
        if not company_id:
            return jsonify({'error': 'No company selected'}), 400
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Get most recent job date for each customer
        query = """
            SELECT 
                c.id,
                c.customer_name,
                MAX(jd.job_date) as last_job_date,
                CURRENT_DATE - MAX(jd.job_date) as days_since
            FROM customers c
            INNER JOIN customer_job_dates jd ON c.id = jd.customer_id
            WHERE c.company_id = %s
            GROUP BY c.id, c.customer_name
            ORDER BY days_since DESC
        """
        
        cur.execute(query, (company_id,))
        customers = cur.fetchall()
        
        cur.close()
        conn.close()
        
        # Convert to JSON-serializable format
        result = []
        for customer in customers:
            result.append({
                'id': customer['id'],
                'name': customer['customer_name'],
                'lastJobDate': customer['last_job_date'].isoformat() if customer['last_job_date'] else None,
                'daysSince': customer['days_since']
            })
        
        return jsonify({
            'success': True,
            'customers': result
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/recency/validate', methods=['POST'])
def validate_recency_upload():
    """Pre-validate a recency upload file before committing data"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        company_id = int(request.form.get('company_id'))

        # Save and clean file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()
        cleaned_path = strip_excel_comments(temp_file.name)
        os.unlink(temp_file.name)

        df = pd.read_excel(cleaned_path, engine='openpyxl')
        os.unlink(cleaned_path)

        if 'Customer' not in df.columns:
            return jsonify({'error': 'Missing Customer column'}), 400

        # Define allowed states per company
        allowed_states = {
            1: ['nc', 'north carolina', 'sc', 'south carolina', 'fl', 'florida'],  # Kleanit Charlotte
            2: ['nc', 'north carolina', 'sc', 'south carolina'],                   # Get a Grip
            3: ['nc', 'north carolina'],                                            # CTS - strict NC only
            4: ['fl', 'florida'],                                                   # Kleanit South Florida
        }

        company_allowed = allowed_states.get(company_id, [])

        # Scan state column
        state_warnings = []
        state_counts = {}
        total_rows = 0

        if 'Service Location State' in df.columns:
            for _, row in df.iterrows():
                customer = row.get('Customer')
                state = row.get('Service Location State')
                if pd.isna(customer) or pd.isna(state):
                    continue
                total_rows += 1
                state_str = str(state).strip().lower()
                state_counts[state_str] = state_counts.get(state_str, 0) + 1
                if state_str not in company_allowed:
                    state_warnings.append(str(state).strip())

        # Check customer name match rate against existing customers
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT customer_name FROM customers WHERE company_id = %s", (company_id,))
        existing = set(r['customer_name'].strip().lower() for r in cur.fetchall())
        cur.close()
        conn.close()

        file_customers = set()
        for _, row in df.iterrows():
            customer = row.get('Customer')
            if not pd.isna(customer):
                file_customers.add(str(customer).strip().lower())

        match_count = sum(1 for c in file_customers if c in existing)
        match_rate = round((match_count / len(file_customers) * 100), 1) if file_customers else 0

        # Build warnings
        warnings = []

        if state_warnings:
            bad_states = list(set(state_warnings))
            warnings.append({
                'type': 'state',
                'message': f"{len(state_warnings)} rows have unexpected states: {', '.join(bad_states)}"
            })

        if len(existing) > 0 and match_rate < 20:
            warnings.append({
                'type': 'customer_match',
                'message': f"Only {match_rate}% of customers in this file match existing {company_id and 'this company'}'s customers ({match_count} of {len(file_customers)}). This may be the wrong company's report."
            })

        return jsonify({
            'valid': len(warnings) == 0,
            'warnings': warnings,
            'summary': {
                'total_rows': total_rows,
                'unique_customers': len(file_customers),
                'match_rate': match_rate,
                'state_counts': state_counts
            }
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


@app.route('/api/recency/batches/<int:company_id>', methods=['GET'])
def get_recency_batches(company_id):
    """Get upload batches grouped by timestamp for a company"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT
                DATE_TRUNC('minute', jd.created_at) as batch_time,
                COUNT(*) as record_count,
                MIN(jd.job_date) as earliest_job,
                MAX(jd.job_date) as latest_job
            FROM customer_job_dates jd
            JOIN customers c ON jd.customer_id = c.id
            WHERE c.company_id = %s
            GROUP BY DATE_TRUNC('minute', jd.created_at)
            ORDER BY batch_time DESC
        """, (company_id,))

        batches = cur.fetchall()
        cur.close()
        conn.close()

        return jsonify({
            'success': True,
            'batches': [{
                'batch_time': b['batch_time'].isoformat(),
                'record_count': b['record_count'],
                'earliest_job': b['earliest_job'].isoformat() if b['earliest_job'] else None,
                'latest_job': b['latest_job'].isoformat() if b['latest_job'] else None
            } for b in batches]
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/recency/clear-batch', methods=['POST'])
def clear_recency_batch():
    """Delete all job dates from a specific upload batch"""
    try:
        data = request.get_json()
        company_id = int(data.get('company_id'))
        batch_time = data.get('batch_time')

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            DELETE FROM customer_job_dates
            WHERE id IN (
                SELECT jd.id
                FROM customer_job_dates jd
                JOIN customers c ON jd.customer_id = c.id
                WHERE c.company_id = %s
                AND DATE_TRUNC('minute', jd.created_at) = DATE_TRUNC('minute', %s::timestamptz)
            )
        """, (company_id, batch_time))

        deleted = cur.rowcount
        conn.commit()
        cur.close()
        conn.close()

        return jsonify({'success': True, 'deleted': deleted})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/recency/stats', methods=['GET'])
def recency_stats():
    """Get statistics about stored job dates by company"""
    try:
        company_id = request.args.get('company_id')
        
        if not company_id:
            return jsonify({'error': 'No company selected'}), 400
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Get stats
        cur.execute("""
            SELECT 
                COUNT(DISTINCT c.id) as total_customers,
                COUNT(jd.id) as total_job_dates,
                MIN(jd.job_date) as earliest_date,
                MAX(jd.job_date) as latest_date
            FROM customers c
            LEFT JOIN customer_job_dates jd ON c.id = jd.customer_id
            WHERE c.company_id = %s
        """, (company_id,))
        
        stats = cur.fetchone()
        
        cur.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'stats': {
                'totalCustomers': stats['total_customers'],
                'totalJobDates': stats['total_job_dates'],
                'earliestDate': stats['earliest_date'].isoformat() if stats['earliest_date'] else None,
                'latestDate': stats['latest_date'].isoformat() if stats['latest_date'] else None
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("\n" + "="*60)
    print("FSM STATEMENT GENERATOR - Starting...")
    print("="*60)
    print(f"Access at: http://localhost:5000")
    print("Press Ctrl+C to stop")
    print("="*60 + "\n")

    app.run(host='0.0.0.0', port=5000, debug=True)
