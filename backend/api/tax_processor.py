"""
Cash-Basis Tax Report Processing
Matches ServiceFusion Tax Report with Transaction Report by Job# to determine
tax liability based on payment collection date (cash basis) per NC requirements.

FIXED: Corrected state vs county tax calculation to use proportions instead of flat percentages
"""

import openpyxl
from datetime import datetime
from collections import defaultdict
import re
from nc_tax_rates import get_tax_breakdown

def parse_date(date_value):
    """Parse various date formats from ServiceFusion exports"""
    if isinstance(date_value, datetime):
        return date_value
    
    if isinstance(date_value, str):
        # Try MM/DD/YYYY format
        try:
            return datetime.strptime(date_value.split()[0], '%m/%d/%Y')
        except:
            pass
        
        # Try MM/DD/YYYY HH:MM am/pm format
        try:
            return datetime.strptime(date_value, '%m/%d/%Y %I:%M %p')
        except:
            pass
    
    return None

def parse_percentage(rate_str):
    """Convert '7.0000%' to 7.0"""
    if isinstance(rate_str, str):
        return float(rate_str.replace('%', ''))
    return float(rate_str)

def process_tax_report(tax_file_path, transaction_file_path, company_id):
    """
    Process both reports and match by Job# to create cash-basis tax report
    
    Returns:
    {
        'success': True,
        'report': {
            'totals': {...},
            'counties': [...]
        }
    }
    """
    
    try:
        # Step 1: Load Tax Report - get tax amounts by Job#
        tax_data = {}  # job_num -> {invoice#, customer, county, tax_rate, tax_amount, total_sales}
        
        wb_tax = openpyxl.load_workbook(tax_file_path)
        ws_tax = wb_tax.active
        
        print(f"Loading tax report: {ws_tax.max_row} rows")
        
        # ServiceFusion Tax Report groups by county - county name appears only on first row of each group
        current_county = 'Unknown'
        
        for row in ws_tax.iter_rows(min_row=2, values_only=True):
            county = row[0]
            invoice_num = row[2]
            customer = row[3]
            job_num = row[4]
            total_sales = row[5]
            tax_rate = row[7]
            tax = row[8]
            
            # Update current county when we see a new county name
            if county and county.strip():
                current_county = county.strip()
            
            # Skip rows without job numbers or zero tax
            if not job_num or not tax or tax == 0:
                continue
            
            # Skip FL customers if this is Kleanit Charlotte (company_id = 1)
            if company_id == '1' and customer and '*FL*' in str(customer).upper():
                continue
            
            # Store tax data indexed by job number, using current county
            tax_data[str(job_num)] = {
                'invoice_num': invoice_num,
                'customer_name': customer,
                'county': current_county,
                'tax_rate': parse_percentage(tax_rate),
                'tax_amount': float(tax),
                'total_sales': float(total_sales) if total_sales else 0
            }
        
        print(f"Loaded {len(tax_data)} tax records from tax report")
        
        # Step 2: Load Transaction Report - get payment dates by Job#
        payment_data = {}  # job_num -> payment_date
        
        wb_trans = openpyxl.load_workbook(transaction_file_path)
        ws_trans = wb_trans.active
        
        print(f"Loading transaction report: {ws_trans.max_row} rows")
        
        for row in ws_trans.iter_rows(min_row=2, values_only=True):
            date_time = row[1]
            job_num = row[2]
            customer_name = row[3]
            trans_type = row[4]
            
            # Only process Payment transactions
            if trans_type != 'Payment' or not job_num:
                continue
            
            # Skip FL customers if this is Kleanit Charlotte (company_id = 1)
            if company_id == '1' and customer_name and '*FL*' in str(customer_name).upper():
                continue
            
            payment_date = parse_date(date_time)
            if not payment_date:
                continue
            
            # Store payment date indexed by job number
            # If multiple payments for same job, use the first one
            if str(job_num) not in payment_data:
                payment_data[str(job_num)] = payment_date
        
        print(f"Loaded {len(payment_data)} payment records from transaction report")
        
        # Step 3: Match tax data with payment dates
        matched_records = []
        unmatched_jobs = []
        
        for job_num, tax_info in tax_data.items():
            if job_num in payment_data:
                payment_date = payment_data[job_num]
                
                matched_records.append({
                    'job_num': job_num,
                    'payment_date': payment_date.isoformat(),
                    'customer_name': tax_info['customer_name'],
                    'county': tax_info['county'],
                    'tax_rate': tax_info['tax_rate'],
                    'tax_amount': tax_info['tax_amount'],
                    'total_sales': tax_info['total_sales']
                })
            else:
                unmatched_jobs.append(job_num)
        
        print(f"Matched {len(matched_records)} records")
        print(f"Unmatched jobs: {len(unmatched_jobs)}")
        
        # Step 4: Group by county and calculate totals
        counties_data = defaultdict(lambda: {
            'name': '',
            'tax_rate': 0,
            'taxable_amount': 0,
            'total_tax': 0,
            'customers': []
        })
        
        for record in matched_records:
            county = record['county']
            
            # Set county info
            if not counties_data[county]['name']:
                counties_data[county]['name'] = county
                counties_data[county]['tax_rate'] = record['tax_rate']
            
            # Add to totals
            counties_data[county]['taxable_amount'] += record['total_sales']
            counties_data[county]['total_tax'] += record['tax_amount']
            
            # Add customer detail
            counties_data[county]['customers'].append({
                'customer_name': record['customer_name'],
                'payment_date': record['payment_date'],
                'total_sales': record['total_sales'],
                'tax': record['tax_amount']
            })
        
        # Step 5: Calculate state vs county tax breakdown using PROPORTIONS
        counties_list = []
        total_state_tax = 0
        total_county_tax = 0
        total_transit_tax = 0
        total_tax = 0
        total_invoices = len(matched_records)
        
        for county_name, county_data in sorted(counties_data.items()):
            total_tax_collected = county_data['total_tax']
            breakdown = get_tax_breakdown(county_name, total_tax_collected)

            counties_list.append({
                'name': county_data['name'],
                'tax_rate': county_data['tax_rate'],
                'taxable_amount': county_data['taxable_amount'],
                'total_tax': county_data['total_tax'],
                'state_tax': breakdown['state'],
                'county_tax': breakdown['county'],
                'transit_tax': breakdown['transit'],
                'customers': sorted(county_data['customers'], key=lambda x: x['customer_name'])
            })

            total_state_tax += breakdown['state']
            total_county_tax += breakdown['county']
            total_transit_tax += breakdown['transit']
            total_tax += county_data['total_tax']
        
        return {
            'success': True,
            'report': {
                'totals': {
                    'total_tax': round(total_tax, 2),
                    'state_tax': round(total_state_tax, 2),
                    'county_tax': round(total_county_tax, 2),
                    'transit_tax': round(total_transit_tax, 2),
                    'invoice_count': total_invoices
                },
                'counties': counties_list
            }
        }
        
    except Exception as e:
        import traceback
        print(f"Error processing tax report: {e}")
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e)
        }
