#!/usr/bin/env python3
"""
FieldKit: ServiceFusion Customer Import
Updated: 2026-06-09
Purpose: Import customers from ServiceFusion Excel export to FieldKit databases.

Changes from original:
  - DB connection updated for Docker (host=db, user=fieldkit, password from env/prompt)
  - Kleanit auto-split: when target is kleanit_charlotte, customers with *FL* in
    their name are automatically routed to fieldkit_kleanit_sf instead.
  - contact_type, accepts_billing, accepts_statements, accepts_general columns
    added to customer_contacts insert (migration 003 columns).
  - Duplicate detection: skips customers whose property_name already exists in
    the target database (prevents double-import).

Usage:
  Run inside the app container:
    docker exec -it fieldkit-phase1-app-1 python3 /app/phase1/fieldkit_phase1/import_sf_customers.py <excel_file> <target_database>

  Target database options:
    getagrip
    kleanit_charlotte   (FL customers auto-routed to kleanit_sf)
    cts
    kleanit_sf

  Example:
    docker exec -it fieldkit-phase1-app-1 python3 /app/phase1/fieldkit_phase1/import_sf_customers.py /tmp/kleanit_customers.xlsx kleanit_charlotte
"""

import sys
import os
import openpyxl
import psycopg2
from datetime import datetime

# ============================================================================
# Database configuration — Docker environment
# ============================================================================

DB_CONFIG = {
    'getagrip':          'fieldkit_getagrip',
    'kleanit_charlotte': 'fieldkit_kleanit_charlotte',
    'cts':               'fieldkit_cts',
    'kleanit_sf':        'fieldkit_kleanit_sf',
}

# Inside the Docker stack, the database container is reachable at host 'db'
DB_HOST = 'db'
DB_PORT = 5432
DB_USER = 'fieldkit'


def connect_db(db_name, password):
    """Connect to a FieldKit database inside the Docker stack."""
    return psycopg2.connect(
        dbname=db_name,
        user=DB_USER,
        password=password,
        host=DB_HOST,
        port=DB_PORT,
    )


# ============================================================================
# Excel parsing
# ============================================================================

def parse_sf_customer_list(filepath):
    """
    Parse a ServiceFusion Customer List Excel export.
    Returns a list of customer dictionaries.
    """
    print(f"Reading {filepath}...")
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    # Find the header row (contains "Customer Name")
    header_row = None
    for row_idx in range(1, 15):
        row_values = [cell.value for cell in sheet[row_idx]]
        if any(val and 'Customer Name' in str(val) for val in row_values):
            header_row = row_idx
            break

    if header_row is None:
        raise ValueError("Could not find header row containing 'Customer Name'")

    headers = [cell.value for cell in sheet[header_row]]
    print(f"Found {len(headers)} columns at row {header_row}")

    customers = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            if header:
                row_data[header] = sheet.cell(row_idx, col_idx).value
        if row_data.get('Customer Name'):
            customers.append(row_data)

    print(f"Parsed {len(customers)} customers from file")
    return customers


# ============================================================================
# FL split logic
# ============================================================================

def is_florida_customer(customer_name):
    """
    Returns True if this customer belongs to Kleanit South Florida.
    SF convention: Florida customers have *FL* anywhere in their name.
    """
    if not customer_name:
        return False
    return '*fl*' in customer_name.lower() or '*FL*' in customer_name


# ============================================================================
# Field helpers
# ============================================================================

def determine_customer_type(customer_name, parent_account):
    name_lower = customer_name.lower() if customer_name else ''

    if any(w in name_lower for w in ['residential', 'homeowner', 'personal']):
        return 'Residential'

    if any(w in name_lower for w in ['contractor', 'construction', 'builder']):
        return 'Contractors'

    if any(w in name_lower for w in [
        'apartment', 'complex', 'village', 'commons', 'place', 'pointe',
        'landing', 'park', 'manor', 'towers', 'ridge', 'crest', 'hills',
        'estates', 'flats', 'lofts', 'residences', 'crossing', 'creek',
        'grove', 'pines', 'oaks', 'lakes', 'court', 'gardens',
    ]):
        return 'Multi Family'

    if parent_account and parent_account.strip():
        return 'Commercial'

    return 'Multi Family'  # Default for Get a Grip / Kleanit


def normalize_state(state_raw):
    if not state_raw:
        return None
    state_map = {
        'North Carolina': 'NC', 'South Carolina': 'SC', 'Georgia': 'GA',
        'Virginia': 'VA', 'Tennessee': 'TN', 'Florida': 'FL',
        'Texas': 'TX', 'New York': 'NY', 'California': 'CA',
    }
    return state_map.get(state_raw, state_raw[:2] if state_raw else None)


def normalize_zip(zip_raw):
    if not zip_raw:
        return None
    return str(zip_raw)[:10]


# ============================================================================
# Management company
# ============================================================================

def find_or_create_management_company(cursor, parent_account, created_by='sf_import'):
    if not parent_account or not parent_account.strip():
        return None

    cursor.execute("""
        SELECT id FROM management_companies
        WHERE name = %s AND deleted_at IS NULL
    """, (parent_account.strip(),))
    result = cursor.fetchone()
    if result:
        return result[0]

    cursor.execute("""
        INSERT INTO management_companies (name, created_by)
        VALUES (%s, %s)
        RETURNING id
    """, (parent_account.strip(), created_by))
    return cursor.fetchone()[0]


# ============================================================================
# Duplicate detection
# ============================================================================

def get_existing_names(cursor):
    """Return a set of all active property_names already in the database."""
    cursor.execute("""
        SELECT property_name FROM customers WHERE deleted_at IS NULL
    """)
    return {row[0] for row in cursor.fetchall()}


# ============================================================================
# Core import
# ============================================================================

def import_customer(cursor, customer_data, created_by='sf_import'):
    """
    Insert a single customer and their contacts.
    Returns (customer_id, contacts_created_count).
    """
    customer_name = (customer_data.get('Customer Name') or '').strip()
    if not customer_name:
        return None, 0

    parent_account = customer_data.get('Parent Account Name')
    account_number = customer_data.get('Account Number')
    is_active      = customer_data.get('Is Active', 'Yes') == 'Yes'
    is_taxable     = customer_data.get('Is Taxable', 'Yes') == 'Yes'
    tax_item       = customer_data.get('Tax Item Name')

    customer_type        = determine_customer_type(customer_name, parent_account)
    management_company_id = find_or_create_management_company(cursor, parent_account, created_by)

    address  = customer_data.get('Primary Service Location Address 1')
    address2 = customer_data.get('Primary Service Location Address 2')
    city     = customer_data.get('Primary Service Location City')
    state    = normalize_state(customer_data.get('Primary Service Location State/Province', ''))
    zip_code = normalize_zip(customer_data.get('Primary Service Location Zip/Postal Code', ''))
    status   = 'Active' if is_active else 'Inactive'

    notes = f"SF Account: {account_number}, Tax: {tax_item if is_taxable else 'Non-taxable'}"

    cursor.execute("""
        INSERT INTO customers (
            property_name, customer_type,
            address, address_2, city, state, zip,
            management_company_id, status, notes, created_by
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (
        customer_name, customer_type,
        address, address2, city, state, zip_code,
        management_company_id, status, notes, created_by,
    ))
    customer_id = cursor.fetchone()[0]

    contacts_created = 0

    # Primary contact
    p_first = customer_data.get('Primary Contact First Name')
    p_last  = customer_data.get('Primary Contact Last Name')
    if p_first or p_last:
        cursor.execute("""
            INSERT INTO customer_contacts (
                customer_id, first_name, last_name, title,
                office_phone, office_email,
                is_primary, contact_type,
                accepts_billing, accepts_statements, accepts_general,
                created_by
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            customer_id,
            p_first or '',
            p_last  or customer_name,
            customer_data.get('Primary Contact Job Title'),
            customer_data.get('Primary Contact Phone 1'),
            customer_data.get('Primary Contact Email 1'),
            True,           # is_primary
            'general',      # contact_type
            True,           # accepts_billing  — first contact gets billing by default
            True,           # accepts_statements
            True,           # accepts_general
            created_by,
        ))
        contacts_created += 1

    # Secondary contact
    s_first = customer_data.get('Secondary Contact First Name')
    s_last  = customer_data.get('Secondary Contact Last Name')
    if s_first or s_last:
        cursor.execute("""
            INSERT INTO customer_contacts (
                customer_id, first_name, last_name, title,
                office_phone, office_email,
                is_primary, contact_type,
                accepts_billing, accepts_statements, accepts_general,
                created_by
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            customer_id,
            s_first or '',
            s_last  or '',
            customer_data.get('Secondary Contact Job Title'),
            customer_data.get('Secondary Contact Phone 1'),
            customer_data.get('Secondary Contact Email 1'),
            False,          # is_primary
            'general',      # contact_type
            False,          # accepts_billing  — secondary defaults off; Michele can enable
            False,          # accepts_statements
            True,           # accepts_general
            created_by,
        ))
        contacts_created += 1

    return customer_id, contacts_created


# ============================================================================
# Per-database import runner
# ============================================================================

def run_import(customers, db_name, password, label):
    """Import a list of customers into a single database. Returns (imported, skipped, contacts)."""
    print(f"\nConnecting to {db_name}...")
    conn = connect_db(db_name, password)
    cursor = conn.cursor()

    existing_names = get_existing_names(cursor)
    print(f"  {len(existing_names)} customers already in database — duplicates will be skipped")

    imported = 0
    skipped_dup = 0
    skipped_err = 0
    total_contacts = 0

    for idx, customer_data in enumerate(customers, 1):
        customer_name = (customer_data.get('Customer Name') or '').strip()

        # Skip duplicates
        if customer_name in existing_names:
            skipped_dup += 1
            continue

        try:
            customer_id, contacts = import_customer(cursor, customer_data)
            if customer_id:
                conn.commit()
                imported += 1
                total_contacts += contacts
                existing_names.add(customer_name)  # prevent intra-batch dupes
                if imported % 100 == 0:
                    print(f"  {imported} imported so far...")
            else:
                skipped_err += 1
        except Exception as e:
            conn.rollback()
            print(f"  Warning: failed to import '{customer_name}': {e}")
            skipped_err += 1

    cursor.close()
    conn.close()

    print(f"\n{'=' * 60}")
    print(f"  {label}")
    print(f"{'=' * 60}")
    print(f"  Imported:          {imported}")
    print(f"  Contacts created:  {total_contacts}")
    print(f"  Skipped (dupes):   {skipped_dup}")
    print(f"  Skipped (errors):  {skipped_err}")

    return imported, skipped_dup + skipped_err, total_contacts


# ============================================================================
# Main
# ============================================================================

def main():
    print("=" * 60)
    print("FieldKit: ServiceFusion Customer Import")
    print("=" * 60)

    if len(sys.argv) < 3:
        print("\nUsage: python3 import_sf_customers.py <excel_file> <target_database>")
        print("\nTarget database options:")
        print("  getagrip          — Get a Grip Charlotte")
        print("  kleanit_charlotte — Kleanit Charlotte (FL customers auto-split to kleanit_sf)")
        print("  cts               — CTS of Raleigh")
        print("  kleanit_sf        — Kleanit South Florida (direct, no split)")
        print("\nExample:")
        print("  python3 import_sf_customers.py /tmp/kleanit.xlsx kleanit_charlotte")
        return 1

    excel_file    = sys.argv[1]
    target_db_key = sys.argv[2]

    if target_db_key not in DB_CONFIG:
        print(f"\nError: '{target_db_key}' is not a valid target.")
        print(f"Valid options: {', '.join(DB_CONFIG.keys())}")
        return 1

    if not os.path.exists(excel_file):
        print(f"\nError: File not found: {excel_file}")
        return 1

    import getpass
    password = getpass.getpass(f"\nPostgreSQL password for user '{DB_USER}': ")

    try:
        customers = parse_sf_customer_list(excel_file)
    except Exception as e:
        print(f"\nError reading Excel file: {e}")
        import traceback
        traceback.print_exc()
        return 1

    # ----------------------------------------------------------------
    # Kleanit auto-split
    # ----------------------------------------------------------------
    if target_db_key == 'kleanit_charlotte':
        charlotte_customers = [c for c in customers if not is_florida_customer(c.get('Customer Name', ''))]
        florida_customers   = [c for c in customers if     is_florida_customer(c.get('Customer Name', ''))]

        print(f"\nKleanit auto-split:")
        print(f"  {len(charlotte_customers)} → fieldkit_kleanit_charlotte")
        print(f"  {len(florida_customers)}   → fieldkit_kleanit_sf (*FL* customers)")

        run_import(charlotte_customers, DB_CONFIG['kleanit_charlotte'], password,
                   "Kleanit Charlotte — Import Summary")
        if florida_customers:
            run_import(florida_customers, DB_CONFIG['kleanit_sf'], password,
                       "Kleanit South Florida — Import Summary")
        else:
            print("\nNo *FL* customers found — skipping kleanit_sf import.")

    # ----------------------------------------------------------------
    # All other targets — single database
    # ----------------------------------------------------------------
    else:
        run_import(customers, DB_CONFIG[target_db_key], password,
                   f"{target_db_key} — Import Summary")

    print("\n✓ Import complete.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
